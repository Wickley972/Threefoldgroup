import csv
from openpyxl import Workbook, load_workbook
import argparse
import logging
from datetime import datetime

logger= logging.getLogger(__name__)
ENCODING = 'utf-8-sig'

# Function to convert a CSV file to an Excel file
def csv_to_excel(input_file, output_file, sheet_name, index, header, delimiter):
    logger.info(input_file, output_file, sheet_name, index, header, delimiter)

    # Load CSV file
    with open(input_file, 'r', encoding=ENCODING) as f:
        reader = csv.reader(f, delimiter=delimiter)
        data = list(reader)

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    if len(sheet_name) > 30:
        sheet_name = sheet_name[:30]
    ws.title = sheet_name

    # Write data to worksheet
    for row in data:
        if header and reader.line_num == 1:
            # If header option is True, write the header row to the worksheet
            ws.append(row)
        elif index:
            # If index option is True, write the row index and data to the worksheet
            ws.append([reader.line_num-1] + row)
        else:
            # Otherwise, write only the data to the worksheet
            ws.append(row)

    # Save Excel file
    wb.save(output_file)

# Function to convert an Excel file to a CSV file
def excel_to_csv(input_file, output_file, delimiter):
    # Load Excel file
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    # Open CSV file for writing
    with open(output_file, 'w', newline='', encoding=ENCODING) as f:
        writer = csv.writer(f, delimiter=delimiter, quoting=1)
        # Write data to CSV file
        #for row in ws.rows:
            #writer.writerow([cell.value for cell in row])
        # Write data to CSV file
        for row in ws.iter_rows(values_only=True):
            formatted_row = []
            for cell_value in row:
                # Format date only if it is a datetime object
                if isinstance(cell_value, datetime) and cell_value.hour == 0 and cell_value.minute == 0 and cell_value.second == 0:
                    formatted_row.append(cell_value.strftime('%Y-%m-%d'))
                else:
                    formatted_row.append(cell_value)
            writer.writerow(formatted_row)
    logger.info('Excel to CSV conversion completed.')


def main(input_file, output_file, sheet, index, header, delimiter, csv2excel=False):
    """parser = argparse.ArgumentParser(description='Convert CSV and Excel files')
    parser.add_argument('input_file', help='Path to input file')
    parser.add_argument('output_file', help='Path to output file')
    parser.add_argument('-s', '--sheet', default='Sheet1', help='Name of the worksheet to create (for CSV to Excel conversion only)')
    #le nom du sheet c'est le nom du fichiers
    parser.add_argument('-i', '--index', action='store_true', help='Include row index in output file (for CSV to Excel conversion only)')
    parser.add_argument('-H', '--header', action='store_true', help='Include header row in output file (for CSV to Excel conversion only)')
    parser.add_argument('-d', '--delimiter', default=',', help='Delimiter to use for CSV file (for Excel to CSV conversion only)')
    parser.add_argument('--excel2csv', action='store_true', help='Convert Excel file to CSV file')
    parser.add_argument('--csv2excel', action='store_true', help='Convert CSV file to Excel file')
    args = parser.parse_args()"""

    logger.info(input_file, output_file, sheet, index, header, delimiter, csv2excel)
    """
    if csv2excel:
        csv_to_excel(input_file, output_file, sheet, index, header, delimiter)
        print('CSV to Excel conversion completed.')
    elif excel2csv:
        excel_to_csv(input_file, output_file, delimiter)
    else:
        print('Please specify either --csv2excel or --excel2csv option.')    
    """
if __name__ == '__main__':
    main()
